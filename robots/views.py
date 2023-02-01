from datetime import datetime

import openpyxl
from django.db.models import Q
from django.forms import model_to_dict
from rest_framework.response import Response
from rest_framework.views import APIView

from robots.models import Order


def write_to_xsl(model, version):
    try:
        wb = openpyxl.load_workbook('robots.xlsx')
    except:
        wb = openpyxl.Workbook()
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)
    if model in wb.sheetnames:
        ws = wb[f'{model}']
        col_b = ws['B']
        for item in col_b:
            if version == item.value:
                ws[f'C{item.row}'].value += 1
                break
            elif item.row == len(col_b):
                count = 1
                ws.append([model, version, count])

    else:
        ws = wb.create_sheet(f'{model}')
        ws['A1'] = 'Модель'
        ws['B1'] = 'Версия'
        ws['C1'] = 'Количество за неделю'
        count = 1
        ws.append([model, version, count])
    wb.save('robots.xlsx')


class RobotsAPIView(APIView):

    def get(self, request):
        orders = Order.objects.all().values()
        return Response({'robots': list(orders)})

    def post(self, request):
        if Order.objects.filter(Q(model=request.data['model']) & Q(version=request.data['version'])):
            write_to_xsl(request.data['model'], request.data['version'])
            return Response({'error': 'Эта модель и версия робота уже существует в системе'})
        new_robot = Order.objects.create(
            serial=f"{request.data['model']}-{request.data['version']}",
            model=request.data['model'],
            version=request.data['version'],
            created=datetime.now()
        )
        write_to_xsl(request.data['model'], request.data['version'])
        return Response({'robot': model_to_dict(new_robot)})
